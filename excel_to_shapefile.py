from openpyxl import load_workbook
from shapely.geometry import Polygon, shape, mapping
import fiona
import itertools
from shapely.ops import unary_union
from sys import exit
from os import remove

# file to convert
excelfile = "ks_2018_grid.xlsx"

# output file, leave off the .shp
outputfile = "kansas_test"

# is the second sheet an attribute table?
attributes = True

# cells have to be the same size, but they don't have to be square
cwidth = 1
cheight = 1

wb = load_workbook(excelfile)

geosheet = wb[wb.sheetnames[0]]

schema = {
    'geometry': 'Polygon',
    'properties': {'id': 'str'}
    }

if attributes:
    attrsheet = wb[wb.sheetnames[1]]
    if attrsheet['A1'].value != "id":
        print ("First column should be named 'id'")
        exit()

    for i in range(len(attrsheet['2'])):
        if attrsheet['2'][i].value not in fiona.FIELD_TYPES_MAP:
            print(attrsheet['2'][i].value + " not in field types map")
            exit()
            
        schema['properties'][attrsheet['1'][i].value] = attrsheet['2'][i].value
            
    attrdata = {}
    
    for i in range(3,attrsheet.max_row+1):
        
        attrdata[attrsheet[str(i)][0].value] = []
        for j in range(1,len(attrsheet['2'])):
            attrdata[attrsheet[str(i)][0].value].append(attrsheet[str(i)][j].value)
            


with fiona.open(outputfile + "_grid.shp", 'w', 'ESRI Shapefile', schema) as map:
    
    curheight = 0
    
    for i in range(geosheet.max_row):
        
        curwidth = 0
        
        for cell in geosheet[str(i+1)]:
            
            if cell.value is not None:
                
                
                    
                polygon = Polygon([[curwidth,curheight],[curwidth+cwidth,curheight],[curwidth+cwidth,curheight-cheight],[curwidth,curheight-cheight]])
                
                outprop = {'id': cell.value}
                
                if attributes:
                    
                    if cell.value not in attrdata:
                    
                        print (cell.value + " not in attribute table")
                        exit(0)
                    
                    for j in range(1,len(attrsheet['1'])):
                        
                        outprop[attrsheet['1'][j].value] = attrdata[cell.value][j-1]
                        
                map.write({
                    'geometry' : mapping(polygon),
                    'properties' : outprop
                    })
                    
                
            curwidth = curwidth + cwidth
            
        curheight = curheight-cheight
        

            
with fiona.open(outputfile + "_grid.shp") as input:

    meta = input.meta
    with fiona.open(outputfile + ".shp", "w", **meta) as output:
        
        e = sorted(input, key=lambda k: k['properties']['id'])
        
        for key, group in itertools.groupby(e, key=lambda x:x['properties']['id']):
            properties, geom = zip(*[(feature['properties'],shape(feature['geometry'])) for feature in group])
            # write the feature, computing the unary_union of the elements in the group with the properties of the first element in the group
            output.write({'geometry': mapping(unary_union(geom)), 'properties': properties[0]})

remove(outputfile + "_grid.shp")
remove(outputfile + "_grid.dbf")
remove(outputfile + "_grid.cpg")
remove(outputfile + "_grid.shx")